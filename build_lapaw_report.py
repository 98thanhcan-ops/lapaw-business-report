from __future__ import annotations

import json
import re
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent
OUTPUT = ROOT / "lapaw_business_report.html"


def num(value) -> float:
    if value in (None, ""):
        return 0.0
    try:
        return float(str(value).replace(",", ""))
    except Exception:
        return 0.0


def clean(value) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.lower() in {"nan", "none", "null"}:
        return ""
    return re.sub(r"\s+", " ", text)


def month_year(name: str) -> tuple[int, int]:
    match = re.search(r"T(\d{1,2})\.(\d{4})", name)
    if not match:
        raise ValueError(f"Cannot parse month/year from {name}")
    return int(match.group(1)), int(match.group(2))


def parse_dt(value):
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value
    text = str(value)
    dayfirst = not bool(re.match(r"^\d{4}-\d{1,2}-\d{1,2}", text))
    parsed = pd.to_datetime(value, errors="coerce", dayfirst=dayfirst)
    if pd.isna(parsed):
        return None
    return parsed.to_pydatetime()


def category(product: str, sku: str) -> str:
    text = f"{product} {sku}".lower()
    if any(x in text for x in ["cát", "lbc", "lbf", "ltf", "tofu", "mix", "datset", "lollipop"]):
        return "Cat litter"
    if any(x in text for x in ["pate", "súp", "sup-", "lpt"]):
        return "Wet food"
    if any(x in text for x in ["hạt", "lkb", "hat"]):
        return "Dry food"
    if any(x in text for x in ["gift", "quà", "khăn", "keomut", "pakeway"]):
        return "Gift / accessory"
    return "Other"


def status_group(status: str) -> str:
    text = clean(status).lower()
    if "hủy" in text:
        return "Đã hủy"
    if any(token in text for token in ["hoàn tất", "hoàn thành", "đã giao", "đã nhận được hàng", "xác nhận đã nhận"]):
        return "Hoàn tất / đã giao"
    if any(token in text for token in ["đang giao", "vận chuyển", "cần vận chuyển"]):
        return "Đang xử lý / vận chuyển"
    return clean(status) or "Không rõ"


def date_iso(dt) -> str:
    return dt.date().isoformat() if dt else ""


def ym(year: int, month: int) -> str:
    return f"{year}-{month:02d}"


def read_shopee() -> tuple[list[dict], list[dict]]:
    orders: list[dict] = []
    lines: list[dict] = []
    for folder in ["Shopee", "Shopee 2026"]:
        for path in sorted((ROOT / folder).glob("*.xlsx")):
            if path.name.startswith("~$"):
                continue
            month, year = month_year(path.name)
            frame = pd.read_excel(path, sheet_name="orders")
            for col in [
                "Số lượng",
                "Giá gốc",
                "Tổng số tiền được người bán trợ giá",
                "Mã giảm giá của Shop",
                "Mã giảm giá của Shopee",
                "Giảm giá từ combo Shopee",
                "Giảm giá từ Combo của Shop",
            ]:
                frame[col] = pd.to_numeric(frame[col], errors="coerce").fillna(0)

            frame["order_id"] = frame["Mã đơn hàng"].astype(str)
            frame["created_at"] = frame["Ngày đặt hàng"].map(parse_dt)
            frame["customer"] = frame["Người Mua"].map(clean)
            frame["status"] = frame["Trạng Thái Đơn Hàng"].map(clean)
            frame["cancelled"] = frame["status"].str.contains("hủy", case=False, na=False)
            frame["line_gross"] = frame["Giá gốc"] * frame["Số lượng"]
            frame["line_pre_voucher"] = frame["line_gross"] - frame["Tổng số tiền được người bán trợ giá"]

            grouped = frame.groupby("order_id", as_index=False).agg(
                created_at=("created_at", "min"),
                customer=("customer", "first"),
                status=("status", "first"),
                cancelled=("cancelled", "max"),
                qty=("Số lượng", "sum"),
                pre_voucher=("line_pre_voucher", "sum"),
                shop_voucher=("Mã giảm giá của Shop", "max"),
                shop_combo=("Giảm giá từ Combo của Shop", "max"),
                province=("Tỉnh/Thành phố", "first"),
                district=("TP / Quận / Huyện", "first"),
            )
            grouped["revenue"] = grouped["pre_voucher"] - grouped["shop_voucher"] - grouped["shop_combo"]
            order_pre = dict(zip(grouped["order_id"], grouped["pre_voucher"]))
            order_discount = dict(zip(grouped["order_id"], grouped["shop_voucher"] + grouped["shop_combo"]))

            for _, row in grouped.iterrows():
                created = row["created_at"]
                oid = f"Shopee::{row['order_id']}"
                orders.append(
                    {
                        "oid": oid,
                        "d": date_iso(created),
                        "ym": ym(year, month),
                        "ch": "Shopee",
                        "st": status_group(row["status"]),
                        "c": 1 if bool(row["cancelled"]) else 0,
                        "rev": round(float(row["revenue"]), 2),
                        "qty": round(float(row["qty"]), 2),
                        "cust": f"Shopee::{clean(row['customer'])}" if clean(row["customer"]) else "",
                        "p": clean(row["province"]),
                        "q": clean(row["district"]),
                        "h": int(created.hour) if created else -1,
                        "w": int(created.weekday()) if created else -1,
                    }
                )

            for _, row in frame.iterrows():
                sku = clean(row.get("SKU phân loại hàng")) or "Không có SKU / chưa phân bổ"
                oid = f"Shopee::{row['order_id']}"
                base = max(float(row["line_pre_voucher"]), 0.0)
                denominator = max(float(order_pre.get(row["order_id"], 0.0)), 0.0)
                allocated_voucher = order_discount.get(row["order_id"], 0.0) * base / denominator if denominator else 0.0
                revenue = float(row["line_pre_voucher"]) - allocated_voucher
                product = clean(row.get("Tên sản phẩm"))
                lines.append(
                    {
                        "d": date_iso(row["created_at"]),
                        "oid": oid,
                        "ym": ym(year, month),
                        "ch": "Shopee",
                        "st": status_group(row["status"]),
                        "c": 1 if bool(row["cancelled"]) else 0,
                        "sku": sku,
                        "cat": category(product, sku),
                        "rev": round(revenue, 2),
                        "qty": round(float(row["Số lượng"]), 2),
                    }
                )
    return orders, lines


def read_tiktok() -> tuple[list[dict], list[dict]]:
    order_map: dict[tuple, dict] = {}
    lines: list[dict] = []
    for folder in ["TikTok Shop", "TikTok Shop 2026"]:
        for path in sorted((ROOT / folder).glob("*.xlsx")):
            if path.name.startswith("~$"):
                continue
            month, year = month_year(path.name)
            workbook = load_workbook(path, read_only=False, data_only=True)
            sheet = workbook.active
            for values in sheet.iter_rows(min_row=3, values_only=True):
                order_id = clean(values[0])
                if not order_id.isdigit():
                    continue
                status = clean(values[1])
                sku = clean(values[6])
                product = clean(values[7])
                qty = num(values[9])
                gross = num(values[12]) or num(values[11]) * qty
                platform_discount = num(values[13])
                seller_discount = num(values[14])
                revenue = gross - seller_discount
                created = parse_dt(values[24])
                customer = clean(values[38])
                province = clean(values[42])
                district = clean(values[43])
                cancelled = "hủy" in status.lower()
                key = (year, month, order_id)
                if key not in order_map:
                    oid = f"TikTok::{year}-{month:02d}::{order_id}"
                    order_map[key] = {
                        "oid": oid,
                        "d": date_iso(created),
                        "ym": ym(year, month),
                        "ch": "TikTok",
                        "st": status_group(status),
                        "c": 1 if cancelled else 0,
                        "rev": 0.0,
                        "qty": 0.0,
                        "cust": f"TikTok::{customer}" if customer else "",
                        "p": province,
                        "q": district,
                        "h": int(created.hour) if created else -1,
                        "w": int(created.weekday()) if created else -1,
                    }
                order = order_map[key]
                order["rev"] += revenue
                order["qty"] += qty
                order["c"] = 1 if order["c"] or cancelled else 0
                lines.append(
                    {
                        "d": date_iso(created),
                        "oid": order["oid"],
                        "ym": ym(year, month),
                        "ch": "TikTok",
                        "st": status_group(status),
                        "c": 1 if cancelled else 0,
                        "sku": sku or "Không có SKU / chưa phân bổ",
                        "cat": category(product, sku),
                        "rev": round(revenue, 2),
                        "qty": round(qty, 2),
                    }
                )
    orders = list(order_map.values())
    for order in orders:
        order["rev"] = round(order["rev"], 2)
        order["qty"] = round(order["qty"], 2)
    return orders, lines


def add_first_dates(orders: list[dict]) -> None:
    first: dict[str, str] = {}
    for row in sorted([r for r in orders if r["cust"] and r["d"]], key=lambda r: r["d"]):
        first.setdefault(row["cust"], row["d"])
    for row in orders:
        row["fd"] = first.get(row["cust"], "")


def compact_lines(lines: list[dict]) -> list[dict]:
    grouped: dict[tuple, dict] = {}
    for row in lines:
        key = (row["oid"], row["d"], row["ym"], row["ch"], row["st"], row["c"], row["sku"], row["cat"])
        if key not in grouped:
            grouped[key] = dict(row)
        else:
            grouped[key]["rev"] += row["rev"]
            grouped[key]["qty"] += row["qty"]
    for row in grouped.values():
        row["rev"] = round(row["rev"], 2)
        row["qty"] = round(row["qty"], 2)
    return list(grouped.values())


def build_report() -> None:
    shopee_orders, shopee_lines = read_shopee()
    tiktok_orders, tiktok_lines = read_tiktok()
    orders = shopee_orders + tiktok_orders
    lines_raw = shopee_lines + tiktok_lines
    lines = compact_lines(lines_raw)
    add_first_dates(orders)
    statuses = sorted({row["st"] for row in orders if row["st"]})
    dates = [row["d"] for row in orders if row["d"]]
    order_fields = ["oid", "d", "ym", "ch", "st", "c", "rev", "qty", "cust", "p", "q", "h", "w", "fd"]
    line_fields = ["oid", "d", "ym", "ch", "st", "c", "sku", "cat", "rev", "qty"]
    data = {
        "orderFields": order_fields,
        "lineFields": line_fields,
        "orders": [[row.get(field, "") for field in order_fields] for row in orders],
        "lines": [[row.get(field, "") for field in line_fields] for row in lines],
        "statuses": statuses,
        "categories": sorted({row["cat"] for row in lines if row.get("cat")}),
        "minDate": min(dates),
        "maxDate": max(dates),
    }

    html = """<!doctype html>
<html lang="vi">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>LaPaw Business Performance Report</title>
  <style>
    :root { --teal:#3f9b8d; --teal-dark:#23786d; --mint:#dff4ec; --gold:#e7bd55; --orange:#e99a5b; --ink:#27302f; --muted:#6f7c7a; --line:#d9e4e1; --panel:#fff; --bg:#f2f5f3; }
    * { box-sizing: border-box; }
    body { margin:0; font-family: Inter, Arial, sans-serif; color:var(--ink); background:var(--bg); }
    body.is-loading .shell { opacity:.45; pointer-events:none; }
    .shell { max-width:1660px; margin:0 auto; padding:20px; }
    .loading { position:fixed; inset:auto 18px 18px auto; background:var(--teal-dark); color:#fff; padding:10px 12px; border-radius:4px; font-weight:800; font-size:13px; box-shadow:0 4px 14px rgba(20,40,36,.18); z-index:20; }
    body:not(.is-loading) .loading { display:none; }
    header { display:flex; justify-content:space-between; align-items:flex-end; gap:24px; margin-bottom:14px; }
    h1 { margin:0; font-size:30px; color:var(--teal-dark); letter-spacing:0; }
    .subtitle,.period,.note { color:var(--muted); font-size:13px; line-height:1.45; }
    .subtitle { margin-top:6px; font-size:14px; }
    .period { text-align:right; }
    .layout { display:grid; grid-template-columns:218px minmax(0,1fr); gap:14px; align-items:start; }
    .filters { position:sticky; top:14px; display:grid; gap:12px; }
    .filter-card { background:#fff; border:1px solid #b7ddd6; border-radius:4px; padding:12px; }
    .filter-card h3 { margin:0 0 8px; font-size:13px; }
    .radio-row { display:flex; gap:8px; align-items:center; margin:7px 0; font-size:13px; }
    select,input[type=date] { width:100%; height:34px; border:1px solid var(--line); border-radius:3px; padding:6px 8px; background:#fff; margin-top:6px; }
    .tabs { display:flex; gap:8px; margin-bottom:14px; }
    .tab { border:1px solid var(--line); background:#fff; color:var(--teal-dark); padding:10px 14px; border-radius:6px; font-weight:700; cursor:pointer; }
    .tab.active { background:var(--teal); color:#fff; border-color:var(--teal); }
    .section { display:none; }
    .section.active { display:block; }
    .grid { display:grid; gap:14px; }
    .kpis { grid-template-columns:repeat(4,minmax(0,1fr)); }
    .two-col { grid-template-columns:1fr; }
    .card { background:var(--panel); border:1px solid var(--line); border-radius:8px; padding:16px; box-shadow:0 1px 2px rgba(20,40,36,.04); }
    .kpi { min-height:112px; display:grid; grid-template-columns:50px 1fr; gap:12px; }
    .icon { width:48px; height:48px; border-radius:8px; display:grid; place-items:center; font-size:18px; font-weight:800; color:#fff; background:var(--teal); }
    .icon.gold { background:var(--gold); } .icon.orange { background:var(--orange); } .icon.dark { background:var(--teal-dark); }
    .kpi-title { font-size:13px; color:var(--muted); font-weight:800; text-transform:uppercase; }
    .kpi-value { font-size:29px; line-height:1.1; margin-top:8px; color:var(--teal-dark); font-weight:800; white-space:nowrap; }
    .kpi-note { margin-top:8px; color:var(--muted); font-size:13px; }
    .kpi-change { display:block; margin-top:4px; font-weight:800; }
    .kpi-change.up { color:var(--teal-dark); }
    .kpi-change.down { color:#c95f55; }
    h2 { margin:0 0 12px; font-size:17px; color:var(--teal-dark); }
    .chart { width:100%; height:320px; }
    .chart-tip { position:fixed; display:none; max-width:240px; background:#173f39; color:#fff; border-radius:4px; padding:9px 10px; font-size:12px; line-height:1.4; box-shadow:0 8px 22px rgba(20,40,36,.22); z-index:30; pointer-events:none; }
    .chart-tip strong { display:block; margin-bottom:4px; color:#dff4ec; }
    .table-head { display:flex; align-items:center; justify-content:flex-end; margin:0 0 8px; }
    .download-btn { border:1px solid var(--line); background:#eef7f4; color:var(--teal-dark); padding:7px 10px; border-radius:4px; font-weight:800; cursor:pointer; font-size:12px; }
    .download-btn:hover { background:#dff4ec; }
    table { width:100%; border-collapse:collapse; font-size:13px; }
    th { background:var(--teal); color:#fff; text-align:right; padding:10px 9px; font-weight:800; position:sticky; top:0; z-index:1; }
    th:first-child,td:first-child { text-align:left; }
    td { padding:9px; text-align:right; border-bottom:1px solid var(--line); }
    tbody tr:nth-child(even) { background:#f7fbf9; }
    .scroll-table { max-height:520px; overflow:auto; border:1px solid var(--line); border-radius:6px; }
    .heatmap-wrap { overflow:auto; }
    .heatmap-wrap table { table-layout:fixed; min-width:1080px; }
    .heatmap-wrap th,.heatmap-wrap td { width:118px; text-align:center; padding:10px 8px; }
    .heatmap-wrap th:first-child,.heatmap-wrap td:first-child { width:110px; min-width:110px; white-space:nowrap; text-align:left; }
    .channel-wrap { overflow:auto; }
    .channel-wrap table { table-layout:fixed; min-width:760px; }
    .channel-wrap th,.channel-wrap td { white-space:nowrap; font-size:12px; padding:9px 8px; }
    .channel-wrap th:first-child,.channel-wrap td:first-child { width:82px; }
    .chart-head { display:flex; align-items:center; justify-content:space-between; gap:12px; margin-bottom:10px; }
    .chart-head h2 { margin:0; }
    .seg { display:flex; gap:0; border:1px solid var(--line); border-radius:6px; overflow:hidden; background:#eef7f4; }
    .seg button { border:0; background:#eef7f4; color:var(--teal-dark); padding:8px 14px; font-weight:800; cursor:pointer; }
    .seg button.active { background:var(--teal); color:#fff; }
    .bar-row { display:grid; grid-template-columns:minmax(130px,230px) minmax(95px,125px) 1fr 78px 62px; gap:10px; align-items:center; margin:10px 0; font-size:13px; }
    .bar-row.with-change { grid-template-columns:minmax(130px,230px) minmax(95px,125px) 1fr 78px 62px 72px 72px; }
    .bar-label { white-space:nowrap; overflow:hidden; text-overflow:ellipsis; }
    .bar-cat { color:var(--muted); white-space:nowrap; overflow:hidden; text-overflow:ellipsis; }
    .bar-track { height:24px; background:#eaf2ef; border-radius:4px; overflow:hidden; }
    .bar-fill { height:100%; background:var(--teal); }
    .bar-value { text-align:right; color:var(--muted); font-weight:700; }
    .matrix { overflow:auto; max-height:620px; border:1px solid var(--line); border-radius:6px; }
    .matrix table { min-width:1500px; }
    .matrix th { min-width:76px; vertical-align:bottom; }
    .matrix th:first-child,.matrix td:first-child { position:sticky; left:0; z-index:2; min-width:190px; }
    .matrix th:first-child { background:var(--teal); }
    .matrix td:first-child { background:#fff; color:var(--ink); font-weight:700; }
    .matrix th:nth-child(2),.matrix td:nth-child(2) { position:sticky; left:190px; z-index:2; min-width:110px; }
    .matrix th:nth-child(2) { background:var(--teal); }
    .matrix td:nth-child(2) { background:#fff; color:var(--muted); }
    .sku-cat { display:block; color:#dff4ec; font-size:10px; font-weight:600; margin-top:3px; }
    .province-row { cursor:pointer; font-weight:800; }
    .province-row td:first-child { color:var(--teal-dark); }
    .district-row td:first-child { padding-left:28px; color:var(--muted); }
    .district-row.hidden { display:none; }
    .toggle { display:inline-block; width:18px; color:var(--teal-dark); font-weight:900; }
    .heat-cell { border-radius:3px; display:block; width:100%; min-width:0; padding:7px 6px; text-align:center; }
    @media (max-width:1180px) { .layout,.kpis { grid-template-columns:1fr; } .filters { position:static; } header { flex-direction:column; align-items:flex-start; } .period { text-align:left; } }
  </style>
</head>
<body>
<main class="shell">
  <header>
    <div><h1>LaPaw Business Performance</h1><div class="subtitle">Shopee + TikTok Shop · FY2025 và MTD 2026 · dashboard có filter</div></div>
    <div class="period">Data source: /Users/nguyencan/Downloads/Lapaw<br>Generated from order exports</div>
  </header>
  <div class="layout">
    <aside class="filters">
      <div class="filter-card">
        <h3>Chọn thời gian</h3>
        <label class="radio-row"><input type="radio" name="period" value="all" checked> Tất cả dữ liệu</label>
        <label class="radio-row"><input type="radio" name="period" value="last30"> 30 ngày qua</label>
        <label class="radio-row"><input type="radio" name="period" value="last7"> 7 ngày qua</label>
        <label class="radio-row"><input type="radio" name="period" value="yesterday"> Hôm qua</label>
        <label class="radio-row"><input type="radio" name="period" value="year"> Năm nay</label>
        <label class="radio-row"><input type="radio" name="period" value="quarter"> Quý này</label>
        <label class="radio-row"><input type="radio" name="period" value="month"> Tháng này</label>
        <label class="radio-row"><input type="radio" name="period" value="week"> Tuần này</label>
        <label class="radio-row"><input type="radio" name="period" value="custom"> Tùy chọn ngày</label>
        <input id="fromDate" type="date"><input id="toDate" type="date">
      </div>
      <div class="filter-card"><h3>Trạng thái đơn hàng</h3><select id="statusFilter"></select></div>
      <div class="filter-card"><h3>Sàn/Shop</h3><select id="channelFilter"><option value="All">All</option><option>Shopee</option><option>TikTok</option></select></div>
      <div class="filter-card"><h3>Danh mục</h3><select id="categoryFilter"></select></div>
    </aside>
    <div>
      <div class="tabs"><button class="tab active" data-tab="business">Tổng doanh số</button><button class="tab" data-tab="product">Doanh số sản phẩm</button></div>
      <section id="business" class="section active">
        <div class="grid kpis">
          <div class="card kpi"><div class="icon">VND</div><div><div class="kpi-title">Doanh số</div><div class="kpi-value" id="kpiRevenue"></div><div class="kpi-note" id="kpiRevenueNote"></div></div></div>
          <div class="card kpi"><div class="icon gold">#</div><div><div class="kpi-title">Số đơn hàng</div><div class="kpi-value" id="kpiOrders"></div><div class="kpi-note" id="kpiAov"></div></div></div>
          <div class="card kpi"><div class="icon orange">ID</div><div><div class="kpi-title">Khách hàng</div><div class="kpi-value" id="kpiCustomers"></div><div class="kpi-note" id="kpiCustomerNote"></div></div></div>
          <div class="card kpi"><div class="icon dark">%</div><div><div class="kpi-title">Tỉ lệ hủy</div><div class="kpi-value" id="kpiCancel"></div><div class="kpi-note" id="kpiCancelNote">Trên tổng đơn theo filter</div></div></div>
        </div>
        <div class="grid two-col" style="margin-top:14px">
          <div class="card"><div class="chart-head"><h2 id="trendTitle">Biến động doanh số và đơn hàng theo tháng</h2><div class="seg"><button class="grain active" data-grain="month">Tháng</button><button class="grain" data-grain="week">Tuần</button><button class="grain" data-grain="day">Ngày</button></div></div><svg id="monthlyChart" class="chart"></svg></div>
          <div class="card"><h2>Chi tiết theo tỉnh/thành và quận/huyện</h2><div class="scroll-table" id="regionTable"></div></div>
        </div>
        <div class="grid two-col" style="margin-top:14px">
          <div class="card"><h2>Doanh thu theo thứ và khung giờ</h2><div class="heatmap-wrap" id="hourHeatmap"></div></div>
          <div class="card"><h2>Doanh số theo kênh</h2><div class="channel-wrap" id="channelTable"></div><div class="note">Shopee DT order-level = Σ(W x Qty - Z) - AF - AK. TikTok DT = M - O. Đơn hủy không tính doanh số.</div></div>
        </div>
        <div class="grid two-col" style="margin-top:14px">
          <div class="card"><h2>Customer type by month</h2><svg id="customerChart" class="chart"></svg><div class="note">New = khách mua lần đầu trong tháng. Returning = khách đã mua trước đó và quay lại trong tháng. Retained = subset của Returning, có mua cả tháng trước.</div></div>
          <div class="card"><h2>Doanh số theo danh mục</h2><div id="categoryTableBusiness"></div></div>
        </div>
      </section>
      <section id="product" class="section">
        <div class="grid kpis">
          <div class="card kpi"><div class="icon">VND</div><div><div class="kpi-title">Doanh số SKU</div><div class="kpi-value" id="productRevenue"></div><div class="kpi-note" id="productRevenueNote">Bao gồm dòng không có SKU / chưa phân bổ</div></div></div>
          <div class="card kpi"><div class="icon gold">SKU</div><div><div class="kpi-title">SKU active</div><div class="kpi-value" id="productSku"></div><div class="kpi-note" id="productPresence"></div></div></div>
          <div class="card kpi"><div class="icon orange">Qty</div><div><div class="kpi-title">Lượng bán</div><div class="kpi-value" id="productQty"></div><div class="kpi-note" id="productQtyNote">Shopee + TikTok</div></div></div>
          <div class="card kpi"><div class="icon dark">Avg</div><div><div class="kpi-title">Giá bán bình quân</div><div class="kpi-value" id="productAsp"></div><div class="kpi-note" id="productAspNote">Doanh số / lượng bán</div></div></div>
        </div>
        <div class="grid two-col" style="margin-top:14px">
          <div class="card"><h2>Top SKU theo doanh số</h2><div id="topSkuRevenue"></div></div>
          <div class="card"><h2>Top SKU theo lượng bán</h2><div id="topSkuQty"></div></div>
        </div>
        <div class="card" style="margin-top:14px"><h2>Top 30 SKU theo 12 tháng gần nhất</h2><div class="matrix" id="skuMatrix"></div></div>
        <div class="card" style="margin-top:14px"><h2>Chi tiết SKU nối 2 kênh</h2><div class="scroll-table" id="productTable"></div><div class="note">Khóa nối: TikTok Seller SKU = Shopee SKU phân loại hàng. SKU không match vẫn giữ riêng theo kênh bán.</div></div>
        <div class="card" style="margin-top:14px"><h2>Raw data doanh số sản phẩm</h2><div class="scroll-table" id="productRawTable"></div><div class="note" id="productRawNote"></div></div>
      </section>
    </div>
  </div>
</main>
<div id="loading" class="loading">Đang xử lý dữ liệu...</div>
<div id="chartTip" class="chart-tip"></div>
<script>
const DATA = __DATA__;
if (Array.isArray(DATA.orders[0])) DATA.orders = DATA.orders.map(row => Object.fromEntries(DATA.orderFields.map((field, idx) => [field, row[idx]])));
if (Array.isArray(DATA.lines[0])) DATA.lines = DATA.lines.map(row => Object.fromEntries(DATA.lineFields.map((field, idx) => [field, row[idx]])));
const WEEKDAYS = ['Thứ hai','Thứ ba','Thứ tư','Thứ năm','Thứ sáu','Thứ bảy','Chủ nhật'];
const HOUR_BUCKETS = [['0-3h',0,3],['3-6h',3,6],['6-9h',6,9],['9-12h',9,12],['12-15h',12,15],['15-18h',15,18],['18-21h',18,21],['21-24h',21,24]];
const moneyShort = value => { const abs = Math.abs(value || 0); if (abs >= 1e9) return (value/1e9).toFixed(2)+'B'; if (abs >= 1e6) return (value/1e6).toFixed(1)+'M'; if (abs >= 1e3) return (value/1e3).toFixed(0)+'K'; return String(Math.round(value || 0)); };
const fmt = value => Math.round(value || 0).toLocaleString('en-US');
const pct = value => ((value || 0) * 100).toFixed(1) + '%';
const esc = value => String(value ?? '').replace(/[&<>"']/g, m => ({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}[m]));
let tableId = 0;
const table = (headers, rows, name='table') => {
  const id = 'tbl' + (++tableId);
  return `<div class="table-head"><button class="download-btn" data-table="${id}" data-name="${esc(name)}">Tải CSV</button></div><table id="${id}"><thead><tr>${headers.map(h=>`<th>${h}</th>`).join('')}</tr></thead><tbody>${rows.map(r=>`<tr>${r.map(c=>`<td>${c}</td>`).join('')}</tr>`).join('')}</tbody></table>`;
};
function csvText(value) { return '"' + String(value ?? '').replace(/"/g, '""') + '"'; }
function downloadRows(headers, rows, name) {
  const body = [headers, ...rows].map(row => row.map(csvText).join(',')).join('\\n');
  const blob = new Blob(['\ufeff' + body], {type:'text/csv;charset=utf-8;'});
  const a = document.createElement('a');
  a.href = URL.createObjectURL(blob);
  a.download = name + '.csv';
  a.click();
  setTimeout(() => URL.revokeObjectURL(a.href), 1000);
}
function downloadTable(button) {
  if (button.dataset.raw === 'product') {
    downloadRows(['Ngày','Tháng','Kênh','Trạng thái','Đơn hủy','SKU','Danh mục','Doanh số','Lượng bán'], rawProductRows, 'lapaw_raw_product_data');
    return;
  }
  const tbl = document.getElementById(button.dataset.table);
  if (!tbl) return;
  const rows = Array.from(tbl.querySelectorAll('tr')).map(tr => Array.from(tr.children).map(td => td.textContent.trim()));
  downloadRows(rows[0] || [], rows.slice(1), button.dataset.name || 'lapaw_table');
}
const toDate = value => new Date(value + 'T00:00:00');
const dateIso = date => new Date(date.getTime() - date.getTimezoneOffset()*60000).toISOString().slice(0,10);
const monthIndex = ym => { const [y,m] = ym.split('-').map(Number); return y*12 + m; };
const monthLabel = idx => { const y = Math.floor((idx - 1)/12); const m = idx - y*12; return y + '-' + String(m).padStart(2,'0'); };
let chartGrain = 'month';
let rawProductRows = [];
function showTip(event, row) {
  const tip = document.getElementById('chartTip');
  if (row.tipType === 'customer') tip.innerHTML = `<strong>${esc(row.label || row.key)}</strong>Số khách: ${fmt(row.rev)}`;
  else if (row.tipType === 'customerMonth') tip.innerHTML = `<strong>${esc(row.label || row.key)}</strong>New: ${fmt(row.rev)}<br>Returning: ${fmt(row.orders)}<br>Retained: ${fmt(row.retained)}`;
  else tip.innerHTML = `<strong>${esc(row.label || row.key)}</strong>Doanh số: ${fmt(row.rev)} VND<br>Số đơn: ${fmt(row.orders)}<br>AOV: ${fmt(row.rev/Math.max(row.orders,1))} VND`;
  tip.style.left = Math.min(event.clientX + 14, window.innerWidth - 260) + 'px';
  tip.style.top = Math.max(12, event.clientY - 14) + 'px';
  tip.style.display = 'block';
}
function hideTip() { document.getElementById('chartTip').style.display = 'none'; }
function groupAdd(map, key, values) { if (!map.has(key)) map.set(key, {}); const row = map.get(key); Object.entries(values).forEach(([k,v]) => row[k] = (row[k] || 0) + v); return row; }
function weekStartIso(value) {
  const d = toDate(value); const diff = (d.getDay()+6)%7; d.setDate(d.getDate()-diff); return dateIso(d);
}
function trendKey(row) {
  if (chartGrain === 'day') return row.d;
  if (chartGrain === 'week') return weekStartIso(row.d);
  return row.ym;
}
function trendLabel(key) {
  if (chartGrain === 'week') return 'W ' + key.slice(5);
  return key;
}
function customerSplit(rows) {
  const counts = new Map();
  rows.filter(r=>r.cust).forEach(r => counts.set(r.cust, (counts.get(r.cust) || 0) + 1));
  let newc = 0, ret = 0;
  counts.forEach(count => { if (count > 1) ret += 1; else newc += 1; });
  return { total: counts.size, newc, ret };
}
function addMonths(ym, delta) {
  const [y, m] = ym.split('-').map(Number);
  return monthLabel(y * 12 + m + delta);
}
function customerTypeByMonth(rows) {
  const monthCustomers = new Map();
  rows.filter(r=>r.cust).forEach(r => {
    if (!monthCustomers.has(r.ym)) monthCustomers.set(r.ym, new Set());
    monthCustomers.get(r.ym).add(r.cust);
  });
  const firstMonth = new Map();
  DATA.orders.filter(r=>!r.c && r.cust).forEach(r => {
    if (!firstMonth.has(r.cust) || r.ym < firstMonth.get(r.cust)) firstMonth.set(r.cust, r.ym);
  });
  return Array.from(monthCustomers.keys()).sort().map(ym => {
    const customers = monthCustomers.get(ym);
    const previous = monthCustomers.get(addMonths(ym, -1)) || new Set();
    let newc = 0, returning = 0, retained = 0;
    customers.forEach(cust => {
      if (firstMonth.get(cust) === ym) newc += 1;
      if (firstMonth.get(cust) < ym) {
        returning += 1;
        if (previous.has(cust)) retained += 1;
      }
    });
    return {ym,total:customers.size,newc,returning,retained};
  });
}
function drawCustomerChart(data) {
  const svg = document.getElementById('customerChart'); const width = svg.clientWidth || 900; const height = svg.clientHeight || 320; const pad = {left:52,right:118,top:24,bottom:54};
  svg.setAttribute('viewBox', `0 0 ${width} ${height}`); svg.innerHTML = '';
  const ns = 'http://www.w3.org/2000/svg';
  const el = (name, attrs) => { const node = document.createElementNS(ns, name); Object.entries(attrs).forEach(([k,v]) => node.setAttribute(k,v)); svg.appendChild(node); return node; };
  const innerW = width-pad.left-pad.right; const innerH = height-pad.top-pad.bottom; const maxValue = Math.max(...data.flatMap(d => [d.newc,d.returning,d.retained]), 1);
  [0,.25,.5,.75,1].forEach(t => { const y = pad.top + innerH - t*innerH; el('line',{x1:pad.left,y1:y,x2:width-pad.right,y2:y,stroke:'#d9e4e1'}); const tx = el('text',{x:pad.left-8,y:y+4,'text-anchor':'end',fill:'#6f7c7a','font-size':11}); tx.textContent = fmt(maxValue*t); });
  const xAt = i => pad.left + (data.length <= 1 ? innerW/2 : i*innerW/(data.length-1));
  const yAt = v => pad.top + innerH - v/maxValue*innerH;
  const series = [
    {key:'newc', label:'New', color:'#6aaee7'},
    {key:'retained', label:'Retained', color:'#6b5fb5'},
    {key:'returning', label:'Returning', color:'#ec7772'},
  ];
  series.forEach(s => {
    el('polyline',{points:data.map((d,i)=>`${xAt(i)},${yAt(d[s.key])}`).join(' '),fill:'none',stroke:s.color,'stroke-width':3,'stroke-linecap':'round','stroke-linejoin':'round'});
    data.forEach((d,i) => { const dot = el('circle',{cx:xAt(i),cy:yAt(d[s.key]),r:4,fill:s.color}); dot.addEventListener('mousemove', event => showTip(event, {key:d.ym,label:`${s.label} · ${d.ym}`,rev:d[s.key],tipType:'customer'})); dot.addEventListener('mouseleave', hideTip); });
  });
  const step = Math.max(1, Math.ceil(data.length / 14));
  data.forEach((d,i) => {
    if (i % step === 0) { const label = el('text',{x:xAt(i),y:height-22,'text-anchor':'end',fill:'#6f7c7a','font-size':10,transform:`rotate(-45 ${xAt(i)} ${height-22})`}); label.textContent = d.ym; }
    const hit = el('rect',{x:xAt(i)-Math.max(12, innerW/Math.max(data.length,1)/2),y:pad.top,width:Math.max(24, innerW/Math.max(data.length,1)),height:innerH,fill:'transparent'});
    hit.addEventListener('mousemove', event => showTip(event, {key:d.ym,label:d.ym,rev:d.newc,orders:d.returning,retained:d.retained,tipType:'customerMonth'}));
    hit.addEventListener('mouseleave', hideTip);
  });
  series.forEach((s,i) => { const x = width - pad.right + 18; const y = pad.top + 14 + i*22; el('circle',{cx:x,cy:y-4,r:4,fill:s.color}); const tx = el('text',{x:x+12,y:y,fill:'#27302f','font-size':12,'font-weight':700}); tx.textContent = s.label; });
}
function currentRange() {
  const selected = document.querySelector('input[name="period"]:checked').value;
  if (selected === 'all') return { from: DATA.minDate, to: DATA.maxDate };
  if (selected === 'custom') return { from: document.getElementById('fromDate').value || DATA.minDate, to: document.getElementById('toDate').value || DATA.maxDate };
  const anchor = toDate(DATA.maxDate); let from = new Date(anchor), to = new Date(anchor);
  if (selected === 'last30') from.setDate(anchor.getDate()-29);
  if (selected === 'last7') from.setDate(anchor.getDate()-6);
  if (selected === 'yesterday') { from.setDate(anchor.getDate()-1); to.setDate(anchor.getDate()-1); }
  if (selected === 'year') from = new Date(anchor.getFullYear(),0,1);
  if (selected === 'quarter') from = new Date(anchor.getFullYear(), Math.floor(anchor.getMonth()/3)*3, 1);
  if (selected === 'month') from = new Date(anchor.getFullYear(), anchor.getMonth(), 1);
  if (selected === 'week') { const diff = (anchor.getDay()+6)%7; from.setDate(anchor.getDate()-diff); }
  return { from: dateIso(from), to: dateIso(to) };
}
function previousRange(range) {
  const from = toDate(range.from); const to = toDate(range.to);
  const days = Math.round((to - from) / 86400000) + 1;
  const prevTo = new Date(from); prevTo.setDate(from.getDate() - 1);
  const prevFrom = new Date(prevTo); prevFrom.setDate(prevTo.getDate() - days + 1);
  return { from: dateIso(prevFrom), to: dateIso(prevTo) };
}
function changeHtml(current, previous, suffix='') {
  if (!isFinite(previous) || previous === 0) return `<span class="kpi-change">vs kỳ trước: --</span>`;
  const change = (current - previous) / Math.abs(previous);
  const cls = change >= 0 ? 'up' : 'down';
  const sign = change >= 0 ? '+' : '';
  return `<span class="kpi-change ${cls}">${sign}${pct(change)} vs kỳ trước${suffix}</span>`;
}
function changeText(current, previous) {
  if (!isFinite(previous) || previous === 0) return '--';
  const change = (current - previous) / Math.abs(previous);
  return (change >= 0 ? '+' : '') + pct(change);
}
function changeClass(current, previous) {
  if (!isFinite(previous) || previous === 0) return '';
  return current >= previous ? 'up' : 'down';
}
function passes(row, range) {
  const status = document.getElementById('statusFilter').value;
  const channel = document.getElementById('channelFilter').value;
  if (row.d && (row.d < range.from || row.d > range.to)) return false;
  if (status !== 'All' && row.st !== status) return false;
  if (channel !== 'All' && row.ch !== channel) return false;
  return true;
}
function filteredForRange(range) {
  const category = document.getElementById('categoryFilter').value;
  const baseOrders = DATA.orders.filter(r => passes(r, range));
  const baseLines = DATA.lines.filter(r => passes(r, range));
  if (category === 'All') return { range, orders: baseOrders, lines: baseLines };
  const lines = baseLines.filter(r => r.cat === category);
  const orderIds = new Set(lines.map(r => r.oid));
  const orders = baseOrders.filter(r => orderIds.has(r.oid));
  return { range, orders, lines, category };
}
function filteredData() {
  const range = currentRange();
  const current = filteredForRange(range);
  current.previous = filteredForRange(previousRange(range));
  return current;
}
function renderBars(id, rows, label, value, formatter=moneyShort, previousMap=null) {
  const max = Math.max(...rows.map(r => r[value]), 1);
  const total = rows.reduce((s, r) => s + (r[value] || 0), 0);
  document.getElementById(id).innerHTML = rows.map(r => {
    const prev = previousMap ? (previousMap.get(r[label]) || {rev:0, qty:0}) : null;
    const changeRev = prev ? `<div class="bar-value kpi-change ${changeClass(r.rev, prev.rev)}">${changeText(r.rev, prev.rev)}</div>` : '';
    const changeQty = prev ? `<div class="bar-value kpi-change ${changeClass(r.qty, prev.qty)}">${changeText(r.qty, prev.qty)}</div>` : '';
    return `<div class="bar-row ${previousMap ? 'with-change' : ''}"><div class="bar-label" title="${esc(r[label])}">${esc(r[label])}</div><div class="bar-cat">${esc(r.cat || '')}</div><div class="bar-track"><div class="bar-fill" style="width:${Math.max(3, r[value]/max*100)}%"></div></div><div class="bar-value">${formatter(r[value])}</div><div class="bar-value">${pct((r[value]||0)/Math.max(total,1))}</div>${changeRev}${changeQty}</div>`;
  }).join('') || '<div class="note">Không có dữ liệu</div>';
}
function drawMonthly(data) {
  const svg = document.getElementById('monthlyChart'); const width = svg.clientWidth || 900; const height = svg.clientHeight || 320; const pad = {left:54,right:40,top:20,bottom:54};
  svg.setAttribute('viewBox', `0 0 ${width} ${height}`); svg.innerHTML = '';
  const maxRevenue = Math.max(...data.map(d => d.rev), 1); const maxOrders = Math.max(...data.map(d => d.orders), 1); const innerW = width-pad.left-pad.right; const innerH = height-pad.top-pad.bottom; const barW = innerW/Math.max(data.length,1)*.62; const ns = 'http://www.w3.org/2000/svg';
  const el = (name, attrs) => { const node = document.createElementNS(ns, name); Object.entries(attrs).forEach(([k,v]) => node.setAttribute(k,v)); svg.appendChild(node); return node; };
  const addTitle = (node, text) => { const title = document.createElementNS(ns, 'title'); title.textContent = text; node.appendChild(title); return node; };
  [0,.25,.5,.75,1].forEach(t => { const y = pad.top + innerH - t*innerH; el('line',{x1:pad.left,y1:y,x2:width-pad.right,y2:y,stroke:'#d9e4e1'}); const tx = el('text',{x:pad.left-8,y:y+4,'text-anchor':'end',fill:'#6f7c7a','font-size':11}); tx.textContent = moneyShort(maxRevenue*t); });
  const points = [];
  const step = Math.max(1, Math.ceil(data.length / 16));
  data.forEach((d,i) => { const x = pad.left + i*innerW/data.length + innerW/data.length/2; const barH = d.rev/maxRevenue*innerH; const bar = addTitle(el('rect',{x:x-barW/2,y:pad.top+innerH-barH,width:barW,height:barH,fill:'#3f9b8d',rx:2}), `${d.label || d.key}\nDoanh số: ${fmt(d.rev)} VND\nSố đơn: ${fmt(d.orders)}`); bar.addEventListener('mousemove', event => showTip(event, d)); bar.addEventListener('mouseleave', hideTip); const y2 = pad.top+innerH-d.orders/maxOrders*innerH; points.push(`${x},${y2}`); if (i % step === 0) { const label = el('text',{x,y:height-22,'text-anchor':'end',fill:'#6f7c7a','font-size':10,transform:`rotate(-45 ${x} ${height-22})`}); label.textContent = d.label || d.key; } });
  el('polyline',{points:points.join(' '),fill:'none',stroke:'#e99a5b','stroke-width':3}); data.forEach((d,i) => { const x = pad.left + i*innerW/data.length + innerW/data.length/2; const y = pad.top+innerH-d.orders/maxOrders*innerH; const dot = addTitle(el('circle',{cx:x,cy:y,r:5,fill:'#e99a5b'}), `${d.label || d.key}\nDoanh số: ${fmt(d.rev)} VND\nSố đơn: ${fmt(d.orders)}`); dot.addEventListener('mousemove', event => showTip(event, d)); dot.addEventListener('mouseleave', hideTip); const hit = el('rect',{x:x-(innerW/data.length/2),y:pad.top,width:innerW/data.length,height:innerH,fill:'transparent'}); hit.addEventListener('mousemove', event => showTip(event, d)); hit.addEventListener('mouseleave', hideTip); });
}
function renderBusiness(f) {
  const allOrders = f.orders; const orders = allOrders.filter(r => !r.c); const revenue = orders.reduce((s,r)=>s+r.rev,0); const qty = orders.reduce((s,r)=>s+r.qty,0);
  const activeCategory = document.getElementById('categoryFilter').value;
  const categoryLines = f.lines.filter(r=>!r.c);
  const categoryRevenue = activeCategory === 'All' ? revenue : categoryLines.reduce((s,r)=>s+r.rev,0);
  const categoryQty = activeCategory === 'All' ? qty : categoryLines.reduce((s,r)=>s+r.qty,0);
  const prevAllOrders = f.previous.orders; const prevOrders = prevAllOrders.filter(r => !r.c); const prevRevenue = prevOrders.reduce((s,r)=>s+r.rev,0); const prevLines = f.previous.lines.filter(r=>!r.c);
  const prevCategoryRevenue = activeCategory === 'All' ? prevRevenue : prevLines.reduce((s,r)=>s+r.rev,0);
  const prevCustomers = customerSplit(prevOrders);
  const prevCancel = prevAllOrders.filter(r=>r.c).length/Math.max(prevAllOrders.length,1);
  const orderRevenue = new Map();
  if (activeCategory === 'All') orders.forEach(r => orderRevenue.set(r.oid, r.rev));
  else categoryLines.forEach(r => orderRevenue.set(r.oid, (orderRevenue.get(r.oid) || 0) + r.rev));
  const customers = customerSplit(orders);
  const cancelRate = allOrders.filter(r=>r.c).length/Math.max(allOrders.length,1);
  document.getElementById('kpiRevenue').textContent = moneyShort(categoryRevenue); document.getElementById('kpiRevenueNote').innerHTML = fmt(categoryRevenue) + ' VND' + changeHtml(categoryRevenue, prevCategoryRevenue); document.getElementById('kpiOrders').textContent = fmt(orders.length); document.getElementById('kpiAov').innerHTML = 'AOV ' + fmt(categoryRevenue/Math.max(orders.length,1)) + ' VND' + changeHtml(orders.length, prevOrders.length); document.getElementById('kpiCustomers').textContent = fmt(customers.total); document.getElementById('kpiCustomerNote').innerHTML = 'New ' + fmt(customers.newc) + ' · Returning ' + fmt(customers.ret) + changeHtml(customers.total, prevCustomers.total); document.getElementById('kpiCancel').textContent = pct(cancelRate); document.getElementById('kpiCancelNote').innerHTML = 'Trên tổng đơn theo filter' + changeHtml(cancelRate, prevCancel);
  document.getElementById('trendTitle').textContent = 'Biến động doanh số và đơn hàng theo ' + (chartGrain === 'day' ? 'ngày' : (chartGrain === 'week' ? 'tuần' : 'tháng'));
  const monthly = new Map(); if (activeCategory === 'All') orders.forEach(r => { const key = trendKey(r); groupAdd(monthly, key, {rev:r.rev, orders:1}); }); else categoryLines.forEach(r => { const key = trendKey(r); groupAdd(monthly, key, {rev:r.rev, orders:0}); }); drawMonthly(Array.from(monthly, ([key,v]) => ({key,label:trendLabel(key),...v})).sort((a,b)=>a.key.localeCompare(b.key)));
  const channel = new Map(); if (activeCategory === 'All') allOrders.forEach(r => groupAdd(channel, r.ch, {total:1,cancelled:r.c?1:0,rev:r.c?0:r.rev,qty:r.c?0:r.qty,orders:r.c?0:1})); else { allOrders.forEach(r => groupAdd(channel, r.ch, {total:1,cancelled:r.c?1:0,rev:0,qty:0,orders:r.c?0:1})); categoryLines.forEach(r => groupAdd(channel, r.ch, {rev:r.rev,qty:r.qty})); } const channelRows = Array.from(channel, ([ch,v]) => ({ch,...v})).sort((a,b)=>b.rev-a.rev); const totalRevenue = channelRows.reduce((s,r)=>s+r.rev,0);
  const prevChannel = new Map(); if (activeCategory === 'All') prevAllOrders.forEach(r => groupAdd(prevChannel, r.ch, {rev:r.c?0:r.rev,qty:r.c?0:r.qty})); else prevLines.forEach(r => groupAdd(prevChannel, r.ch, {rev:r.rev,qty:r.qty}));
  document.getElementById('channelTable').innerHTML = table(['Kênh','Doanh số','% đổi DT','Số đơn','Lượng bán','% đổi SL','AOV','Tỉ lệ hủy','Tỷ trọng'], channelRows.map(r => { const prev = prevChannel.get(r.ch) || {rev:0,qty:0}; return [esc(r.ch),fmt(r.rev),`<span class="kpi-change ${changeClass(r.rev, prev.rev)}">${changeText(r.rev, prev.rev)}</span>`,fmt(r.orders),fmt(r.qty),`<span class="kpi-change ${changeClass(r.qty, prev.qty)}">${changeText(r.qty, prev.qty)}</span>`,fmt(r.rev/Math.max(r.orders,1)),pct(r.cancelled/Math.max(r.total,1)),pct(r.rev/Math.max(totalRevenue,1))]; }), 'doanh_so_theo_kenh');
  drawCustomerChart(customerTypeByMonth(orders));
  const provinceMap = new Map();
  orders.forEach(r => {
    const p = r.p || 'Không rõ'; const q = r.q || 'Không rõ';
    const rev = orderRevenue.get(r.oid) || 0;
    if (!provinceMap.has(p)) provinceMap.set(p, {p, rev:0, orders:0, districts:new Map()});
    const province = provinceMap.get(p);
    province.rev += rev; province.orders += 1;
    groupAdd(province.districts, q, {rev, orders:1});
  });
  const provinces = Array.from(provinceMap.values()).sort((a,b)=>b.rev-a.rev).slice(0,40);
  const regionTotal = provinces.reduce((s,p)=>s+p.rev,0);
  const regionRows = [];
  provinces.forEach((p, idx) => {
    const id = 'p' + idx;
    regionRows.push(`<tr class="province-row" data-province="${id}"><td><span class="toggle">+</span>${esc(p.p)}</td><td></td><td>${fmt(p.rev)}</td><td>${fmt(p.orders)}</td><td>${fmt(p.rev/Math.max(p.orders,1))}</td><td>${pct(p.rev/Math.max(regionTotal,1))}</td></tr>`);
    Array.from(p.districts, ([q,v]) => ({q,...v})).sort((a,b)=>b.rev-a.rev).slice(0,20).forEach(d => {
      regionRows.push(`<tr class="district-row hidden" data-parent="${id}"><td>${esc(d.q)}</td><td></td><td>${fmt(d.rev)}</td><td>${fmt(d.orders)}</td><td>${fmt(d.rev/Math.max(d.orders,1))}</td><td>${pct(d.rev/Math.max(p.rev,1))}</td></tr>`);
    });
  });
  const regionTableId = 'tbl' + (++tableId);
  document.getElementById('regionTable').innerHTML = `<div class="table-head"><button class="download-btn" data-table="${regionTableId}" data-name="chi_tiet_tinh_quan">Tải CSV</button></div><table id="${regionTableId}"><thead><tr><th>Tỉnh/Thành phố</th><th>Quận/Huyện</th><th>Doanh số</th><th>Số đơn</th><th>AOV</th><th>% share</th></tr></thead><tbody>${regionRows.join('')}</tbody></table>`;
  document.querySelectorAll('.province-row').forEach(row => row.addEventListener('click', () => {
    const id = row.dataset.province; const open = row.classList.toggle('open');
    row.querySelector('.toggle').textContent = open ? '-' : '+';
    document.querySelectorAll(`.district-row[data-parent="${id}"]`).forEach(child => child.classList.toggle('hidden', !open));
  }));
  const orderById = new Map(orders.map(r => [r.oid, r]));
  const heat = Array.from({length:HOUR_BUCKETS.length}, () => Array(7).fill(0)); orders.forEach(r => { const b = HOUR_BUCKETS.findIndex(x => r.h>=x[1] && r.h<x[2]); if (b>=0 && r.w>=0) heat[b][r.w] += activeCategory === 'All' ? r.rev : 0; }); if (activeCategory !== 'All') categoryLines.forEach(r => { const order = orderById.get(r.oid); if (!order) return; const b = HOUR_BUCKETS.findIndex(x => order.h>=x[1] && order.h<x[2]); if (b>=0 && order.w>=0) heat[b][order.w] += r.rev; });
  const rowTotals = heat.map(row => row.reduce((s,v)=>s+v,0));
  const colTotals = WEEKDAYS.map((_, col) => heat.reduce((s,row)=>s+row[col],0));
  const grandTotal = rowTotals.reduce((s,v)=>s+v,0);
  const maxHeat = Math.max(...heat.flat(), ...rowTotals, ...colTotals, grandTotal, 1);
  const heatCell = (v, strong=false) => `<span style="display:block;background:rgba(63,155,141,${v ? 0.12+0.72*v/maxHeat : 0});padding:7px 8px;border-radius:3px;${strong ? 'font-weight:800' : ''}">${moneyShort(v)}</span>`;
  const heatRows = HOUR_BUCKETS.map((b,i) => [`<strong>${b[0]}</strong>`, ...heat[i].map(v => heatCell(v)), heatCell(rowTotals[i], true)]);
  heatRows.push(['<strong>Tổng</strong>', ...colTotals.map(v => heatCell(v, true)), heatCell(grandTotal, true)]);
  document.getElementById('hourHeatmap').innerHTML = table(['Nhóm giờ',...WEEKDAYS,'Tổng'], heatRows, 'doanh_thu_theo_thu_khung_gio');
  const catMap = new Map(); const lines = f.lines.filter(r=>!r.c);
  lines.forEach(r => groupAdd(catMap, r.cat || 'Other', {rev:r.rev,qty:r.qty}));
  const catTotal = Array.from(catMap.values()).reduce((s,r)=>s+(r.rev||0),0);
  document.getElementById('categoryTableBusiness').innerHTML = table(['Danh mục','Doanh số','Lượng bán','% share'], Array.from(catMap, ([cat,v]) => ({cat,...v})).sort((a,b)=>b.rev-a.rev).map(r => [esc(r.cat),fmt(r.rev),fmt(r.qty),pct(r.rev/Math.max(catTotal,1))]), 'doanh_so_theo_danh_muc');
}
function renderProduct(f) {
  const lines = f.lines.filter(r=>!r.c); const skuMap = new Map(); const catMap = new Map();
  lines.forEach(r => { const item = skuMap.get(r.sku) || {sku:r.sku,cat:r.cat||'Other',rev:0,qty:0,shopee:0,tiktok:0,channels:new Set()}; item.rev += r.rev; item.qty += r.qty; item.channels.add(r.ch); if (r.ch==='Shopee') item.shopee += r.rev; if (r.ch==='TikTok') item.tiktok += r.rev; skuMap.set(r.sku,item); groupAdd(catMap, item.cat, {rev:r.rev,qty:r.qty}); });
  const skus = Array.from(skuMap.values()).sort((a,b)=>b.rev-a.rev); const revenue = skus.reduce((s,r)=>s+r.rev,0); const qty = skus.reduce((s,r)=>s+r.qty,0); const both = skus.filter(r=>r.channels.has('Shopee')&&r.channels.has('TikTok')).length; const shOnly = skus.filter(r=>r.channels.has('Shopee')&&!r.channels.has('TikTok')).length; const ttOnly = skus.filter(r=>r.channels.has('TikTok')&&!r.channels.has('Shopee')).length;
  const prevLines = f.previous.lines.filter(r=>!r.c); const prevSkuSet = new Set(prevLines.map(r=>r.sku)); const prevRevenue = prevLines.reduce((s,r)=>s+r.rev,0); const prevQty = prevLines.reduce((s,r)=>s+r.qty,0); const asp = revenue/Math.max(qty,1); const prevAsp = prevRevenue/Math.max(prevQty,1);
  const prevSkuMap = new Map();
  prevLines.forEach(r => { const item = prevSkuMap.get(r.sku) || {rev:0, qty:0}; item.rev += r.rev; item.qty += r.qty; prevSkuMap.set(r.sku, item); });
  document.getElementById('productRevenue').textContent = moneyShort(revenue); document.getElementById('productRevenueNote').innerHTML = 'Bao gồm dòng không có SKU / chưa phân bổ' + changeHtml(revenue, prevRevenue); document.getElementById('productSku').textContent = fmt(skus.length); document.getElementById('productPresence').innerHTML = `Both ${fmt(both)} · Shopee-only ${fmt(shOnly)} · TikTok-only ${fmt(ttOnly)}` + changeHtml(skus.length, prevSkuSet.size); document.getElementById('productQty').textContent = fmt(qty); document.getElementById('productQtyNote').innerHTML = 'Shopee + TikTok' + changeHtml(qty, prevQty); document.getElementById('productAsp').textContent = fmt(asp); document.getElementById('productAspNote').innerHTML = 'Doanh số / lượng bán' + changeHtml(asp, prevAsp);
  renderBars('topSkuRevenue', skus.slice(0,12), 'sku', 'rev', moneyShort, prevSkuMap); renderBars('topSkuQty', [...skus].sort((a,b)=>b.qty-a.qty).slice(0,12), 'sku', 'qty', fmt, prevSkuMap);
  document.getElementById('productTable').innerHTML = table(['SKU','Danh mục','Tổng DT','Shopee DT','TikTok DT','Lượng bán','Kênh','% share'], skus.slice(0,100).map(r => [esc(r.sku),esc(r.cat),fmt(r.rev),fmt(r.shopee),fmt(r.tiktok),fmt(r.qty),r.channels.size===2?'Both':(r.channels.has('Shopee')?'Shopee only':'TikTok only'),pct(r.rev/Math.max(revenue,1))]), 'chi_tiet_sku_noi_2_kenh');
  const top30 = skus.slice(0,30); const maxMonth = Math.max(...lines.map(r=>monthIndex(r.ym)), monthIndex(DATA.maxDate.slice(0,7))); const months = Array.from({length:12}, (_,i)=>monthLabel(maxMonth-11+i)); const monthSku = new Map(); lines.forEach(r => { if (months.includes(r.ym)) monthSku.set(r.ym+'||'+r.sku, (monthSku.get(r.ym+'||'+r.sku)||0)+r.rev); });
  const maxMatrix = Math.max(...months.flatMap(m => top30.map(s => monthSku.get(m+'||'+s.sku)||0)), 1);
  const matrixHeaders = ['SKU', 'Danh mục', ...months, 'Tổng', '% share'];
  const matrixRows = top30.map(s => {
    const total = months.reduce((sum, m) => sum + (monthSku.get(m+'||'+s.sku)||0), 0);
    return [esc(s.sku), esc(s.cat), ...months.map(m => {
      const value = monthSku.get(m+'||'+s.sku)||0;
    const alpha = value ? 0.10 + 0.78 * value / maxMatrix : 0;
    const style = value ? `background:rgba(63,155,141,${alpha})` : 'background:#fff';
    return `<span class="heat-cell" style="${style}">${fmt(value)}</span>`;
    }), `<strong>${fmt(total)}</strong>`, pct(total/Math.max(revenue,1))];
  });
  document.getElementById('skuMatrix').innerHTML = table(matrixHeaders, matrixRows, 'top_30_sku_12_thang');
  rawProductRows = lines.sort((a,b)=>b.d.localeCompare(a.d) || b.rev-a.rev).map(r => [r.d, r.ym, r.ch, r.st, r.c ? 'Có' : 'Không', r.sku, r.cat || '', Math.round(r.rev || 0), r.qty || 0]);
  const preview = rawProductRows.slice(0,500).map(r => [esc(r[0]),esc(r[1]),esc(r[2]),esc(r[3]),esc(r[4]),esc(r[5]),esc(r[6]),fmt(r[7]),fmt(r[8])]);
  document.getElementById('productRawTable').innerHTML = `<div class="table-head"><button class="download-btn" data-raw="product">Tải CSV toàn bộ raw</button></div>` + table(['Ngày','Tháng','Kênh','Trạng thái','Đơn hủy','SKU','Danh mục','Doanh số','Lượng bán'], preview, 'raw_product_preview');
  document.getElementById('productRawNote').textContent = `Đang preview ${fmt(Math.min(rawProductRows.length,500))}/${fmt(rawProductRows.length)} dòng theo filter. Nút tải CSV toàn bộ raw xuất đầy đủ tất cả dòng.`;
}
function updateDashboard() {
  document.body.classList.add('is-loading');
  setTimeout(() => {
    tableId = 0;
    const f = filteredData();
    const active = document.querySelector('.section.active')?.id || 'business';
    if (active === 'business') renderBusiness(f);
    if (active === 'product') renderProduct(f);
    document.body.classList.remove('is-loading');
  }, 0);
}
function init() {
  document.getElementById('statusFilter').innerHTML = '<option value="All">All</option>' + DATA.statuses.map(s => `<option value="${esc(s)}">${esc(s)}</option>`).join('');
  document.getElementById('categoryFilter').innerHTML = '<option value="All">All</option>' + DATA.categories.map(s => `<option value="${esc(s)}">${esc(s)}</option>`).join('');
  document.getElementById('fromDate').value = DATA.minDate; document.getElementById('toDate').value = DATA.maxDate;
  document.querySelectorAll('input[name="period"], #fromDate, #toDate, #statusFilter, #channelFilter, #categoryFilter').forEach(el => el.addEventListener('change', updateDashboard));
  document.querySelectorAll('.grain').forEach(button => button.addEventListener('click', () => { document.querySelectorAll('.grain').forEach(x=>x.classList.remove('active')); button.classList.add('active'); chartGrain = button.dataset.grain; updateDashboard(); }));
  document.querySelectorAll('.tab').forEach(button => button.addEventListener('click', () => { document.querySelectorAll('.tab').forEach(x=>x.classList.remove('active')); document.querySelectorAll('.section').forEach(x=>x.classList.remove('active')); button.classList.add('active'); document.getElementById(button.dataset.tab).classList.add('active'); updateDashboard(); }));
  document.addEventListener('click', event => { const button = event.target.closest('.download-btn'); if (button) downloadTable(button); });
  document.body.classList.add('is-loading');
  window.addEventListener('resize', updateDashboard); updateDashboard();
}
init();
</script>
</body>
</html>
""".replace("__DATA__", json.dumps(data, ensure_ascii=False))
    OUTPUT.write_text(html, encoding="utf-8")
    print(OUTPUT)
    print(json.dumps({"orders": len(orders), "lines": len(lines), "minDate": data["minDate"], "maxDate": data["maxDate"]}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    build_report()
